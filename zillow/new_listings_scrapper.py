import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import subprocess
import json
import csv
import traceback
from datetime import datetime
import re
import random
import os
import time
import urllib.parse
import functools
import builtins
import requests
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
)
logger = logging.getLogger(__name__)

_original_print = builtins.print
def print(*args, **kwargs):
    msg = ' '.join(str(a) for a in args)
    _original_print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | {msg}", flush=True)
builtins.print = print

# -- Configuration --------------------------------------------------------------
ZAPIER_WEBHOOK_URL = ""
ZAPIER_DELAY       = 0.6

URL_TRACKING_FILE   = ""
URL_TRACKING_COLUMN = ""
# ------------------------------------------------------------------------------

# -- Xvfb for headless display -------------------------------------------------
try:
    subprocess.Popen(
        ['Xvfb', ':99', '-screen', '0', '1920x1080x24'],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
    )
    time.sleep(2)
    os.environ['DISPLAY'] = ':99'
except Exception:
    pass


def _get_chrome_version():
    try:
        result = subprocess.run(['google-chrome', '--version'], capture_output=True, text=True)
        m = re.search(r'(\d+)\.', result.stdout)
        return int(m.group(1)) if m else None
    except Exception:
        return None


# -- URL Tracking helpers -------------------------------------------------------
def load_scraped_urls() -> set:
    """
    Load previously scraped ZPIDs from the tracking file (.xlsx or .csv).
    Uses ZPID-based matching to avoid URL format mismatches.
    """
    import openpyxl

    def extract_zpid(url):
        m = re.search(r'/(\d+)_zpid', str(url))
        return m.group(1) if m else None

    if not os.path.exists(URL_TRACKING_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([URL_TRACKING_COLUMN, 'state', 'scraped_at'])
        wb.save(URL_TRACKING_FILE)
        print(f" ?? Created new URL tracking file: {URL_TRACKING_FILE}")
        return set()

    zpids = set()
    try:
        ext = os.path.splitext(URL_TRACKING_FILE)[1].lower()
        if ext in ('.xlsx', '.xlsm', '.xls'):
            wb = openpyxl.load_workbook(URL_TRACKING_FILE)
            ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            col_idx = headers.index(URL_TRACKING_COLUMN) if URL_TRACKING_COLUMN in headers else 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                zpid = extract_zpid(row[col_idx])
                if zpid:
                    zpids.add(zpid)
        else:
            with open(URL_TRACKING_FILE, 'r', encoding='utf-8') as f:
                for row in csv.DictReader(f):
                    zpid = extract_zpid(row.get(URL_TRACKING_COLUMN, ''))
                    if zpid:
                        zpids.add(zpid)
        print(f" ?? Loaded {len(zpids)} previously scraped ZPIDs from {URL_TRACKING_FILE}")
    except Exception as e:
        print(f" ?? Could not read URL tracking file: {e}. Starting fresh.")
    return zpids


def append_urls_to_file(new_entries: list):
    """Append newly scraped entries to the tracking file."""
    import openpyxl
    if not new_entries:
        return
    try:
        ext = os.path.splitext(URL_TRACKING_FILE)[1].lower()
        if ext in ('.xlsx', '.xlsm', '.xls'):
            if os.path.exists(URL_TRACKING_FILE):
                wb = openpyxl.load_workbook(URL_TRACKING_FILE)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append([URL_TRACKING_COLUMN, 'state', 'scraped_at'])
            for entry in new_entries:
                ws.append([entry.get(URL_TRACKING_COLUMN, ''),
                            entry.get('state', ''),
                            entry.get('scraped_at', '')])
            wb.save(URL_TRACKING_FILE)
        else:
            file_exists = os.path.exists(URL_TRACKING_FILE)
            with open(URL_TRACKING_FILE, 'a', newline='', encoding='utf-8') as f:
                w = csv.DictWriter(f, fieldnames=[URL_TRACKING_COLUMN, 'state', 'scraped_at'])
                if not file_exists:
                    w.writeheader()
                w.writerows(new_entries)
        print(f" ?? Appended {len(new_entries)} new URLs to {URL_TRACKING_FILE}")
    except Exception as e:
        print(f" ? Could not append to URL tracking file: {e}")


# -- Zapier sender --------------------------------------------------------------

def send_to_zapier(records: list, retries: int = 3) -> bool:
    """Send records to Zapier one-by-one via GET query params."""
    if not records:
        return True

    success_count = 0
    fail_count    = 0
    print(f" ?? Sending {len(records)} records to Zapier...")

    for idx, record in enumerate(records, 1):
        payload = {k: str(v).strip() if v is not None else '' for k, v in record.items()}
        sent = False
        for attempt in range(1, retries + 1):
            try:
                resp = requests.get(ZAPIER_WEBHOOK_URL, params=payload, timeout=30)
                if resp.status_code == 200:
                    print(f"   ? [{idx}/{len(records)}] status=200  {resp.text[:60]}")
                    sent = True
                    break
                else:
                    print(f"   ?? Record {idx}: HTTP {resp.status_code}")
            except requests.exceptions.RequestException as e:
                print(f"   ? Record {idx} attempt {attempt}/{retries}: {e}")
            if attempt < retries:
                time.sleep(attempt * 3)

        if sent:
            success_count += 1
        else:
            fail_count += 1
            print(f"   ? Gave up on record {idx}: {record.get('listing_url', 'unknown')}")
        time.sleep(ZAPIER_DELAY)

    print(f" ? Zapier send done â€” {success_count} ok / {fail_count} failed")
    return fail_count == 0


# -- Main Scraper ---------------------------------------------------------------

class ZillowActiveListingsScraper:

    def __init__(self, headless=True, output_dir='./zillow_data_active', max_captcha_retries=3):
        self.headless            = headless
        self.output_dir          = output_dir
        self.max_captcha_retries = max_captcha_retries

        os.makedirs(output_dir, exist_ok=True)

        self.scraped_urls        = load_scraped_urls()
        self.captcha_blocked_file = os.path.join(output_dir, 'captcha_blocked_urls.json')
        self.captcha_blocked_urls = self._load_captcha_blocked_urls()

        self.user_agents = [
            'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
            'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
        ]

        self.states = {
            'AL': 'Alabama',        'AK': 'Alaska',
            'AZ': 'Arizona',        'AR': 'Arkansas',
            'CA': 'California',     'CO': 'Colorado',
            'CT': 'Connecticut',    'DE': 'Delaware',
            'FL': 'Florida',        'GA': 'Georgia',
            'HI': 'Hawaii',         'ID': 'Idaho',
            'IL': 'Illinois',       'IN': 'Indiana',
            'IA': 'Iowa',           'KS': 'Kansas',
            'KY': 'Kentucky',       'LA': 'Louisiana',
            'ME': 'Maine',          'MD': 'Maryland',
            'MA': 'Massachusetts',  'MI': 'Michigan',
            'MN': 'Minnesota',      'MS': 'Mississippi',
            'MO': 'Missouri',       'MT': 'Montana',
            'NE': 'Nebraska',       'NV': 'Nevada',
            'NH': 'New Hampshire',  'NJ': 'New Jersey',
            'NM': 'New Mexico',     'NY': 'New York',
            'NC': 'North Carolina', 'ND': 'North Dakota',
            'OH': 'Ohio',           'OK': 'Oklahoma',
            'OR': 'Oregon',         'PA': 'Pennsylvania',
            'RI': 'Rhode Island',   'SC': 'South Carolina',
            'SD': 'South Dakota',   'TN': 'Tennessee',
            'TX': 'Texas',          'UT': 'Utah',
            'VT': 'Vermont',        'VA': 'Virginia',
            'WA': 'Washington',     'WV': 'West Virginia',
            'WI': 'Wisconsin',      'WY': 'Wyoming',
        }

        self.state_region_ids = {
            'AL': 5,  'AK': 2,  'AZ': 3,  'AR': 4,  'CA': 9,  'CO': 6,
            'CT': 7,  'DE': 8,  'FL': 11, 'GA': 13, 'HI': 15, 'ID': 16,
            'IL': 17, 'IN': 18, 'IA': 19, 'KS': 20, 'KY': 21, 'LA': 22,
            'ME': 23, 'MD': 24, 'MA': 25, 'MI': 26, 'MN': 27, 'MS': 28,
            'MO': 29, 'MT': 30, 'NE': 31, 'NV': 32, 'NH': 33, 'NJ': 34,
            'NM': 35, 'NY': 36, 'NC': 37, 'ND': 38, 'OH': 39, 'OK': 40,
            'OR': 41, 'PA': 42, 'RI': 44, 'SC': 45, 'SD': 46, 'TN': 47,
            'TX': 48, 'UT': 49, 'VT': 50, 'VA': 51, 'WA': 53, 'WV': 54,
            'WI': 55, 'WY': 56,
        }

    # --- Persistence ------------------------------------------------------

    def _load_captcha_blocked_urls(self):
        if os.path.exists(self.captcha_blocked_file):
            try:
                with open(self.captcha_blocked_file, 'r') as f:
                    return json.load(f)
            except Exception:
                return []
        return []

    def _save_captcha_blocked_url(self, url, state_code):
        entry = {
            'url':       url,
            'state':     state_code,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'zpid':      self._extract_zpid_from_url(url),
        }
        self.captcha_blocked_urls.append(entry)
        with open(self.captcha_blocked_file, 'w') as f:
            json.dump(self.captcha_blocked_urls, f, indent=2)

    # --- URL generation ----------------------------------------------------

    def generate_state_url(self, state_code):
        """
        Generate a Zillow search URL for active listings in the given state.
        Uses the state code in the path (e.g., /wi/) and a minimal filter set
        that has been proven to work (based on the second, working scraper).
        """
        base_url = f"https://www.zillow.com/{state_code.lower()}/"
        params = {
            "isMapVisible": True,
            "filterState": {
                "price": {"min": 7000000, "max": None},
                "mp": {"min": 35000, "max": None},
                "land": {"value": False},
                "manu": {"value": False},
                "sort": {"value": "globalrelevanceex"}
            },
            "isListVisible": True
        }
        encoded = urllib.parse.quote(json.dumps(params, separators=(',', ':')))
        return f"{base_url}?searchQueryState={encoded}"

    # --- Helpers -----------------------------------------------------------

    @staticmethod
    def _extract_zpid_from_url(url):
        match = re.search(r'/(\d+)_zpid', url)
        return match.group(1) if match else None

    def _parse_date(self, date_string):
        if not date_string:
            return None
        try:
            date_string = re.sub(
                r'^(Listed|Updated|Sold|Listing updated|Zillow last checked):?\s*',
                '', date_string, flags=re.IGNORECASE)
            for fmt in ('%B %d, %Y at %I:%M%p', '%B %d, %Y', '%m/%d/%Y', '%Y-%m-%d'):
                try:
                    return datetime.strptime(date_string.strip(), fmt).strftime('%Y-%m-%d')
                except ValueError:
                    continue
            date_match = re.search(r'(\w+ \d+, \d{4})', date_string)
            if date_match:
                return datetime.strptime(date_match.group(1), '%B %d, %Y').strftime('%Y-%m-%d')
        except Exception:
            pass
        return None

    def _random_delay(self, min_seconds=2, max_seconds=5):
        time.sleep(random.uniform(min_seconds, max_seconds))

    # --- CAPTCHA handling --------------------------------------------------

    def _is_captcha_showing(self, driver):
        indicators = ['press & hold', 'press and hold', 'before we continue',
                      'confirm you are', 'not a bot']
        try:
            page_source = driver.page_source.lower()
            for indicator in indicators:
                if indicator in page_source:
                    return True
        except Exception:
            pass
        try:
            elems = driver.find_elements(By.ID, 'px-captcha')
            if elems and elems[0].is_displayed():
                return True
        except Exception:
            pass
        return False

    def _find_captcha_button(self, driver):
        print("   ? Searching for CAPTCHA button...")
        for by, selector, label in [
            (By.ID, 'px-captcha', '#px-captcha'),
            (By.XPATH, "//button[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'press')]", 'button:press'),
            (By.XPATH, "//*[@role='button'][contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'press')]", 'role=button:press'),
        ]:
            try:
                elems = driver.find_elements(by, selector)
                for elem in elems:
                    if elem.is_displayed():
                        loc  = elem.location
                        size = elem.size
                        if size['width'] > 10 and size['height'] > 10:
                            text = ''
                            try: text = elem.text[:40]
                            except Exception: pass
                            print(f"   ? Found ({label}): '{text}' at ({loc['x']},{loc['y']})")
                            return elem, {'x': loc['x'], 'y': loc['y'],
                                          'width': size['width'], 'height': size['height']}
            except Exception:
                continue

        for tag in ('button', '*[@role="button"]', 'div'):
            try:
                elems = driver.find_elements(By.XPATH, f"//{tag}")
                for elem in elems:
                    try:
                        if not elem.is_displayed(): continue
                        text = (elem.text or '').lower()
                        if 'press' in text or 'hold' in text:
                            loc  = elem.location
                            size = elem.size
                            if size['width'] > 10 and size['height'] > 10:
                                print(f"   ? Found (generic {tag}): '{text[:40]}'")
                                return elem, {'x': loc['x'], 'y': loc['y'],
                                              'width': size['width'], 'height': size['height']}
                    except Exception:
                        continue
            except Exception:
                continue

        try:
            iframes = driver.find_elements(By.TAG_NAME, 'iframe')
            if iframes:
                print(f"   ? Checking {len(iframes)} iframes...")
                for idx, iframe in enumerate(iframes):
                    try:
                        driver.switch_to.frame(iframe)
                        for by, selector in [(By.ID, 'px-captcha'),
                                             (By.XPATH, "//*[@role='button']"),
                                             (By.TAG_NAME, 'button')]:
                            for elem in driver.find_elements(by, selector):
                                try:
                                    if not elem.is_displayed(): continue
                                    text = (elem.text or '').lower()
                                    if 'press' in text or 'hold' in text or selector == 'px-captcha':
                                        loc  = elem.location
                                        size = elem.size
                                        if size['width'] > 10 and size['height'] > 10:
                                            print(f"   ? Found in iframe {idx}: '{text[:40]}'")
                                            return elem, {'x': loc['x'], 'y': loc['y'],
                                                          'width': size['width'], 'height': size['height']}
                                except Exception:
                                    continue
                        driver.switch_to.default_content()
                    except Exception:
                        try: driver.switch_to.default_content()
                        except Exception: pass
        except Exception:
            try: driver.switch_to.default_content()
            except Exception: pass

        return None, None

    def _solve_press_and_hold_captcha(self, driver):
        print("   ?? Attempting to solve Press & Hold CAPTCHA...")
        try:
            time.sleep(2)
            button, box = self._find_captcha_button(driver)
            if not button or not box:
                print("   ? CAPTCHA button not found")
                return False

            cx = box['x'] + box['width']  / 2
            cy = box['y'] + box['height'] / 2

            steps   = random.randint(8, 15)
            start_x = random.uniform(100, 500)
            start_y = random.uniform(100, 400)
            for step in range(steps):
                t  = (step + 1) / steps
                t  = t * t * (3 - 2 * t)
                mx = start_x + (cx - start_x) * t + random.uniform(-2, 2)
                my = start_y + (cy - start_y) * t + random.uniform(-2, 2)
                driver.execute_cdp_cmd('Input.dispatchMouseEvent',
                                       {'type': 'mouseMoved', 'x': mx, 'y': my, 'pointerType': 'mouse'})
                time.sleep(random.uniform(0.01, 0.04))

            time.sleep(random.uniform(0.3, 0.7))
            hold_duration = random.uniform(12, 16)
            print(f"   ? Pressing and HOLDING for {hold_duration:.1f}s...")
            driver.execute_cdp_cmd('Input.dispatchMouseEvent',
                                   {'type': 'mousePressed', 'x': cx, 'y': cy,
                                    'button': 'left', 'clickCount': 1, 'pointerType': 'mouse'})

            start_time = time.time()
            last_log   = 0
            while True:
                time.sleep(0.5)
                elapsed = time.time() - start_time
                if int(elapsed) > last_log:
                    last_log = int(elapsed)
                    print(f"   ? Holding... {int(elapsed)}s / {int(hold_duration)}s")
                if elapsed >= hold_duration:
                    break

            driver.execute_cdp_cmd('Input.dispatchMouseEvent',
                                   {'type': 'mouseReleased', 'x': cx, 'y': cy,
                                    'button': 'left', 'clickCount': 1, 'pointerType': 'mouse'})
            print(f"   ? Released after {hold_duration:.1f}s")

            for wait_sec in range(15):
                time.sleep(1)
                if not self._is_captcha_showing(driver):
                    print(f"   ? CAPTCHA solved! (cleared after {wait_sec+1}s)")
                    return True

            print("   ? CAPTCHA still showing after hold")
            return False

        except Exception as e:
            print(f"   ? Solver error: {e}")
            traceback.print_exc()
            return False

    def _check_and_handle_captcha(self, driver, url, state_code):
        if not self._is_captcha_showing(driver):
            return False

        print("\n ?? CAPTCHA detected!")
        for attempt in range(5):
            print(f"\n   Solve attempt {attempt+1}/5")
            if self._solve_press_and_hold_captcha(driver):
                print(" ? CAPTCHA solved automatically!")
                return True
            if not self._is_captcha_showing(driver):
                print(" ? CAPTCHA gone!")
                return True
            time.sleep(random.uniform(3, 6) + attempt * 2)

        if self._is_captcha_showing(driver):
            print(f"\n{'='*60}")
            print("?? Please solve CAPTCHA manually (5 min timeout)")
            print(f"{'='*60}\n")
            for i in range(300):
                time.sleep(1)
                if not self._is_captcha_showing(driver):
                    print(" ? Solved manually! Continuing...")
                    time.sleep(2)
                    return True
                if i % 30 == 29:
                    print(f"   ? Still waiting... ({(i+1)//60}m {(i+1)%60}s elapsed)")

        self._save_captcha_blocked_url(url, state_code)
        raise Exception("CAPTCHA could not be resolved")

    # --- Scrolling / link collection ---------------------------------------

    def _scroll_to_load_all_listings(self, driver, max_scrolls=40):
        print(" ?? Loading all listings...")
        previous_count  = 0
        no_change_count = 0
        for scroll in range(max_scrolls):
            try:
                current_count = len(driver.find_elements(By.CSS_SELECTOR,
                    'article, [class*="StyledCard"], [class*="property-card"], '
                    '[data-test="property-card"], a[href*="_zpid"]'))
                print(f"   Scroll {scroll+1}: {current_count} items found", end='\r')
                if current_count == previous_count:
                    no_change_count += 1
                    if no_change_count >= 5:
                        print(f"\n   ? Stable at {current_count} items")
                        break
                else:
                    no_change_count = 0
                previous_count = current_count
                driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
                self._random_delay(0.5, 1.5)
            except Exception as e:
                print(f"\n ?? Scroll error: {e}")
                break
        return previous_count

    def _extract_property_links(self, driver):
        property_links = {}
        selectors = [
            'article a[href*="/homedetails/"]',
            'a[data-test="property-card-link"]',
            'a[href*="_zpid"]',
            '[class*="StyledCard"] a[href*="_zpid"]',
            '[class*="property-card"] a[href*="_zpid"]',
            'ul[class*="photo-cards"] a[href*="_zpid"]',
            'div[class*="list-card"] a[href*="_zpid"]',
        ]
        for selector in selectors:
            try:
                for link in driver.find_elements(By.CSS_SELECTOR, selector):
                    try:
                        href = link.get_attribute('href')
                        if href and '_zpid' in href and '/homedetails/' in href:
                            full_url = href if href.startswith('http') else f"https://www.zillow.com{href}"
                            zpid = self._extract_zpid_from_url(full_url)
                            if zpid:
                                property_links[zpid] = full_url
                    except Exception:
                        continue
            except Exception:
                continue

        # Fallback
        if not property_links:
            try:
                for link in driver.find_elements(By.XPATH, '//a[contains(@href,"_zpid")]'):
                    try:
                        href = link.get_attribute('href')
                        if href and '/homedetails/' in href:
                            full_url = href if href.startswith('http') else f"https://www.zillow.com{href}"
                            zpid = self._extract_zpid_from_url(full_url)
                            if zpid:
                                property_links[zpid] = full_url
                    except Exception:
                        continue
            except Exception:
                pass
        return property_links

    def _go_to_next_page(self, driver):
        for selector in ('a[title="Next page"]', 'a[aria-label="Next page"]',
                         'button[aria-label="Next page"]', 'a[rel="next"]'):
            try:
                for btn in driver.find_elements(By.CSS_SELECTOR, selector):
                    if not btn.is_displayed():
                        continue
                    classes       = (btn.get_attribute('class') or '').lower()
                    aria_disabled = btn.get_attribute('aria-disabled') or ''
                    if 'disabled' in classes or aria_disabled == 'true':
                        print(" ?? Next button found but disabled")
                        return False
                    try:
                        if not btn.is_enabled():
                            return False
                    except Exception:
                        pass
                    print(" ? Found active 'Next' button")
                    driver.execute_script("arguments[0].click();", btn)
                    time.sleep(2)
                    WebDriverWait(driver, 15).until(
                        lambda d: d.execute_script("return document.readyState") == "complete")
                    self._random_delay(2, 4)
                    return True
            except Exception:
                continue
        return False

    def _collect_all_listing_links(self, driver, state_code):
        new_links         = {}
        current_page      = 1
        max_pages         = 30
        pages_with_no_new = 0

        print(f"\n {'='*60}")
        print(f" COLLECTING LINKS FOR {state_code}")
        print(f" ZPIDs already tracked: {len(self.scraped_urls)}")
        print(f" {'='*60}\n")

        while current_page <= max_pages:
            print(f" ?? Page {current_page}")
            try:
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR,
                        'article, div[id*="search-page-list"], a[href*="_zpid"], '
                        'ul[class*="photo-cards"], [data-test="property-card"]')))
                self._random_delay(2, 3)
            except Exception:
                print(f" ?? Timed out waiting for listings on page {current_page}")
                break

            self._scroll_to_load_all_listings(driver, max_scrolls=40)
            page_links = self._extract_property_links(driver)

            if not page_links:
                print(f" ?? 0 links found on page {current_page}")

            genuinely_new   = 0
            already_tracked = 0
            duplicate_run   = 0

            for zpid, url in page_links.items():
                if zpid in self.scraped_urls:
                    already_tracked += 1
                elif zpid in new_links:
                    duplicate_run += 1
                else:
                    new_links[zpid] = url
                    genuinely_new  += 1

            print(f" ? Page {current_page}: {len(page_links)} found on page | "
                  f"{genuinely_new} new | "
                  f"{already_tracked} already tracked | "
                  f"{duplicate_run} duplicates")
            print(f" ?? Running total of NEW listings to scrape: {len(new_links)}")

            if genuinely_new == 0:
                pages_with_no_new += 1
                if pages_with_no_new >= 2:
                    print(f" ? Stopping â€” no new listings for {pages_with_no_new} consecutive pages\n")
                    break
            else:
                pages_with_no_new = 0

            print(" ? Checking for next page...")
            if not self._go_to_next_page(driver):
                print(" ? No more pages\n")
                break

            current_page += 1
            self._random_delay(3, 6)

        print(f" {'='*60}")
        print(f" ? TOTAL NEW LINKS FOR {state_code}: {len(new_links)}")
        print(f" {'='*60}\n")
        return list(new_links.values())

    # --- Data extraction ----------------------------------------------------

    def _extract_listing_data(self, driver, state_code):
        data = {
            'source_platform': 'Zillow',
            'listing_id':      None,
            'listing_url':     driver.current_url,
            'address_full':    None,
            'city':            None,
            'state':           state_code,   
            'zip':             None,
            'price':           None,
            'beds':            None,
            'baths':           None,
            'sqft':            None,
            'price_per_sqft':  None,
            'lot_size':        None,
            'property_type':   None,
            'status':          'For Sale',
            'listed_date':     None,
            'agent_name':      None,
            'agent_email':     None,
            'agent_phone':     None,
            'agent_company':   None,
            'mls_id':          None,
            'days_on_zillow':  None,
            'views':           None,
            'saves':           None,
            'year_built':      None,
            'description':     None,
            'scraped_at':      datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }

        try:
            WebDriverWait(driver, 10).until(
                lambda d: d.execute_script("return document.readyState") == "complete")
            time.sleep(random.uniform(1, 2))

            data['listing_id'] = self._extract_zpid_from_url(driver.current_url)
            page_text = driver.find_element(By.TAG_NAME, 'body').text

            # Beds / Baths / Sqft â€” combined pattern first, then individual
            for pattern in (
                r'(\d+)\s*beds?\s+([\d.]+)\s*baths?\s+([\d,]+)\s*sqft',
                r'(\d+)\s*bd\s+([\d.]+)\s*ba\s+([\d,]+)\s*sqft',
            ):
                m = re.search(pattern, page_text, re.IGNORECASE)
                if m:
                    data['beds']  = m.group(1)
                    data['baths'] = m.group(2)
                    data['sqft']  = m.group(3).replace(',', '')
                    break

            if not data['beds']:
                m = re.search(r'(\d+)beds?\b', page_text, re.IGNORECASE)
                if m: data['beds'] = m.group(1)
            if not data['baths']:
                m = re.search(r'([\d.]+)baths?\b', page_text, re.IGNORECASE)
                if m: data['baths'] = m.group(1)
            if not data['sqft']:
                m = re.search(r'([\d,]+)sqft', page_text, re.IGNORECASE)
                if m: data['sqft'] = m.group(1).replace(',', '')

            # Price
            cleaned = re.sub(r'Price cut:\s*\$[0-9,.]+[KMB]?', '', page_text, flags=re.IGNORECASE)
            m = re.search(r'\$([0-9,]{6,})', cleaned)
            if m:
                data['price'] = m.group(1).replace(',', '')

            # Address â€” city and zip only, state always from state_code
            for selector in ('h1[data-test="property-address"]', 'h1'):
                try:
                    for elem in driver.find_elements(By.CSS_SELECTOR, selector):
                        text = elem.text
                        if ',' in text:
                            data['address_full'] = text.strip()
                            parts = text.split(',')
                            if len(parts) >= 2:
                                data['city'] = parts[-2].strip()
                                last = parts[-1].strip().split()
                                if len(last) >= 2:
                                    # state always forced to state_code (not last[0])
                                    data['zip'] = last[1]
                            break
                except Exception:
                    continue

            # Price per sqft
            m = re.search(r'\$([0-9,]+)/sqft', page_text)
            if m: data['price_per_sqft'] = m.group(1).replace(',', '')

            # Days on Zillow
            m = re.search(r'(\d+)\s+days?\s+on\s+Zillow', page_text, re.IGNORECASE)
            if m: data['days_on_zillow'] = m.group(1)

            # Views
            m = re.search(r'([\d,]+)\s+views?', page_text, re.IGNORECASE)
            if m: data['views'] = m.group(1).replace(',', '')

            # Saves
            m = re.search(r'([\d,]+)\s+saves?', page_text, re.IGNORECASE)
            if m: data['saves'] = m.group(1).replace(',', '')

            # Agent info
            self._extract_agent_info(data, page_text)

            # MLS
            m = re.search(r'MLS#?:?\s*([\w-]+)', page_text, re.IGNORECASE)
            if m: data['mls_id'] = m.group(1)

            # Listed date
            for pattern in (
                r'Listed:?\s*([A-Za-z]+\s+\d+,\s+\d{4})',
                r'Listing\s+updated:?\s*([A-Za-z]+\s+\d+,\s+\d{4})',
                r'(\d{1,2}/\d{1,2}/\d{4})',
            ):
                m = re.search(pattern, page_text, re.IGNORECASE)
                if m:
                    data['listed_date'] = self._parse_date(m.group(1))
                    if data['listed_date']:
                        break

            # Property type
            for ptype in ('Single Family', 'Multi Family', 'MultiFamily', 'Condo',
                          'Townhouse', 'Apartment', 'SingleFamily'):
                if ptype in page_text:
                    data['property_type'] = ptype.replace('MultiFamily', 'Multi Family').replace('SingleFamily', 'Single Family')
                    break

            # Lot size
            m = re.search(r'([\d,]+\.?\d*)\s*(acres?|sqft)', page_text, re.IGNORECASE)
            if m: data['lot_size'] = f"{m.group(1)} {m.group(2)}"

            # Year built
            m = re.search(r'(?:Year\s+built:?\s*|Built\s+in\s+)(19|20\d{2})', page_text, re.IGNORECASE)
            if m: data['year_built'] = m.group(1)

            # Description
            try:
                desc_elems = driver.find_elements(By.CSS_SELECTOR, 'div[data-testid="description"]')
                if desc_elems:
                    data['description'] = desc_elems[0].text.replace('Show more', '').strip()[:500]
            except Exception:
                pass

        except Exception as e:
            print(f" ?? Extraction error: {e}")

        return data

    def _extract_agent_info(self, data, page_text):
        # Pattern 1: Name + email + company + phone
        m = re.search(
            r'Listed by:\s*([A-Z][a-z]+(?:\s+[A-Z][a-z]+)+)\s+'
            r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}),\s*'
            r'([A-Za-z\s,.\-&]+?)\s+(\d{3}-\d{3}-\d{4})',
            page_text, re.IGNORECASE)
        if m:
            data['agent_name']    = m.group(1).strip()
            data['agent_email']   = m.group(2).strip()
            data['agent_company'] = m.group(3).strip()
            data['agent_phone']   = m.group(4)
            return

        # Pattern 2: Name + PHONE:xxx + company
        m = re.search(
            r'Listed by:\s*([A-Z][a-z]+(?:\s+[A-Z][a-z]+)+)\s+PHONE:(\d{3}-\d{3}-\d{4}),\s*'
            r'(.+?)(?:Source:|MLS|$)',
            page_text, re.IGNORECASE | re.DOTALL)
        if m:
            data['agent_name']    = m.group(1).strip()
            data['agent_phone']   = m.group(2).strip()
            data['agent_company'] = ' '.join(m.group(3).split()).rstrip(',')
            return

        # Pattern 3: multi-line name + phone + company
        m = re.search(
            r'Listed by:\s*\n\s*\n\s*([A-Z][a-z]+(?:\s+[A-Z][a-z]+)+)\s+(\d{3}-\d{3}-\d{4}),\s*\n\s*\n\s*([^\n,]+)',
            page_text, re.IGNORECASE)
        if m:
            data['agent_name']    = m.group(1).strip()
            data['agent_phone']   = m.group(2).strip()
            data['agent_company'] = self._clean_company(m.group(3).strip().rstrip(','))
            return

        # Fallbacks
        if not data['agent_email']:
            m = re.search(r'\b([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})\b', page_text)
            if m: data['agent_email'] = m.group(1)

        if not data['agent_phone']:
            m = re.search(r'(\d{3}[-.]?\d{3}[-.]?\d{4})', page_text)
            if m: data['agent_phone'] = self._normalize_phone(m.group(1))

        if data['agent_phone'] and not data['agent_name']:
            phone_pos = page_text.find(data['agent_phone'])
            if phone_pos > 0:
                context = page_text[max(0, phone_pos-200):phone_pos]
                names = re.findall(r'\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,3})\b', context)
                exclude = {'Contact Information','Listing Updated','Listing Agent','Source',
                           'Mobile','Office','Phone','Call','Email','Real Estate','Press Release',
                           'For Sale','Days On','Listed By','Co-Listing Agent'}
                company_kw = {'Realty','Properties','Group','Company','Agency','Associates',
                              'Partners','International','LLC','Inc','Team','Homes','Estate','Compass','Sotheby'}
                for name in reversed(names):
                    if not any(e in name for e in exclude) and not any(k in name for k in company_kw):
                        data['agent_name'] = name
                        break

        if not data['agent_company']:
            m = re.search(
                r"([A-Z][A-Za-z\s&',.]+"
                r"(?:Realty|Real Estate|Properties|Sotheby's International Realty|Group|Company|Agency|LLC|Inc|Compass))",
                page_text)
            if m:
                company = re.sub(r'\s*(?:Source|MLS|Contact).*$', '', m.group(1)).strip()
                data['agent_company'] = self._clean_company(company)

    @staticmethod
    def _clean_company(company):
        if not company:
            return company
        if "Zillow Group is committed" in company:
            return None
        return company

    @staticmethod
    def _normalize_phone(phone):
        clean = phone.replace('-', '').replace('.', '').replace(' ', '')
        if len(clean) == 10 and clean.isdigit():
            return f"{clean[:3]}-{clean[3:6]}-{clean[6:]}"
        return phone

    # --- Scraping loop ------------------------------------------------------

    def _scrape_listings(self, driver, property_links, state_code):
        # Returns a tuple of (possibly new driver, state_data, url_entries)
        # because we may need to restart the browser during the loop.
        state_data          = []
        url_entries         = []
        total               = len(property_links)
        captcha_retry_count = 0

        from selenium.common import exceptions as selenium_exceptions

        print(f"\n {'='*60}")
        print(f" SCRAPING {total} NEW LISTINGS FOR {state_code}")
        print(f" {'='*60}\n")

        for i, url in enumerate(property_links, 1):
            retry_count = 0
            while True:
                try:
                    print(f" [{i}/{total}] {url}")
                    driver.get(url)
                    self._random_delay(2, 4)

                    try:
                        self._check_and_handle_captcha(driver, url, state_code)
                        captcha_retry_count = 0
                    except Exception as captcha_error:
                        captcha_retry_count += 1
                        print(f" ? CAPTCHA error: {captcha_error}")
                        if captcha_retry_count >= self.max_captcha_retries:
                            print(" ?? Extended break (60-120s)...")
                            time.sleep(random.uniform(60, 120))
                            captcha_retry_count = 0
                        # if we hit a captcha, skip this listing and move on
                        break

                    listing_data = self._extract_listing_data(driver, state_code)

                    if listing_data and listing_data.get('address_full'):
                        # Validate listing belongs to the state being scraped
                        url_match  = f'-{state_code}-' in listing_data['listing_url'].upper() or \
                                     f'/{state_code}/' in listing_data['listing_url'].upper()
                        addr       = listing_data.get('address_full', '')
                        addr_match = f', {state_code} ' in addr or addr.endswith(f', {state_code}')

                        if not url_match and not addr_match:
                            print(f" ?? SKIPPED â€” wrong state (scraping {state_code}): {addr}")
                            break

                        state_data.append(listing_data)
                        zpid = self._extract_zpid_from_url(listing_data['listing_url'])
                        if zpid:
                            self.scraped_urls.add(zpid)
                        url_entries.append({
                            URL_TRACKING_COLUMN: listing_data['listing_url'],
                            'state':             state_code,
                            'scraped_at':        listing_data['scraped_at'],
                        })

                        print(f" ? {listing_data.get('address_full','N/A')} â€” ${listing_data.get('price','N/A')}")
                        print(f"   Agent      : {listing_data.get('agent_name','None')} | {listing_data.get('agent_company','None')}")
                        print(f"   Phone      : {listing_data.get('agent_phone','None')} | Email: {listing_data.get('agent_email','None')}")
                        print(f"   Days/Views : {listing_data.get('days_on_zillow','None')} days | "
                              f"{listing_data.get('views','None')} views | {listing_data.get('saves','None')} saves")
                    else:
                        print(" ?? No data extracted (no address found)")

                    if i % 5 == 0 and i < total:
                        print(f"\n ?? Break after {i} listings...\n")
                        self._random_delay(8, 15)
                    else:
                        self._random_delay(3, 6)

                    # successfully processed this url, move to next
                    break

                except Exception as e:
                    # If the browser connection died, restart it and retry once
                    msg = str(e)
                    print(f" ? Error scraping {url}: {e}")
                    if retry_count < 2 and (
                            'Connection refused' in msg or
                            'RemoteDisconnected' in msg or
                            isinstance(e, selenium_exceptions.WebDriverException)):
                        retry_count += 1
                        print("   ? browser appears dead; restarting and retrying...")
                        try:
                            driver.quit()
                        except Exception:
                            pass
                        driver = self._create_driver()
                        # on a fresh driver we need to navigate to the url again
                        continue
                    # otherwise give up on this listing
                    break

        print(f"\n {'='*60}")
        print(f" ? SCRAPING DONE â€” {len(state_data)} valid listings for {state_code}")
        print(f" {'='*60}\n")

        return driver, state_data, url_entries

    # --- Driver -------------------------------------------------------------

    def _create_driver(self):
        options = uc.ChromeOptions()
        if self.headless:
            options.add_argument('--headless=new')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--start-maximized')
        options.add_argument('--disable-infobars')
        options.add_argument('--disable-background-timer-throttling')
        options.add_argument('--disable-backgrounding-occluded-windows')
        options.add_argument('--disable-renderer-backgrounding')
        options.add_argument('--window-size=1920,1080')
        options.add_argument(f'--user-agent={random.choice(self.user_agents)}')
        ver = _get_chrome_version()
        print(f" Chrome version: {ver}")
        driver = uc.Chrome(options=options, use_subprocess=True, version_main=ver)
        time.sleep(random.uniform(1, 3))
        return driver

    # --- State orchestration -------------------------------------------------

    def scrape_state(self, state_code, driver):
        print(f"\n{'#'*70}")
        print(f"# STARTING: {state_code} â€” {self.states[state_code]}")
        print(f"{'#'*70}\n")

        search_url = self.generate_state_url(state_code)
        print(f" URL: {search_url}\n")

        driver.get('https://www.zillow.com')
        time.sleep(2)
        driver.get(search_url)
        self._random_delay(3, 5)
        self._check_and_handle_captcha(driver, search_url, state_code)

        new_links = self._collect_all_listing_links(driver, state_code)

        if not new_links:
            print(f" ? No new listings to scrape for {state_code}\n")
            # nothing to do; propagate current driver back to caller
            return driver

        # _scrape_listings may restart the browser and return a new driver instance
        driver, state_data, url_entries = self._scrape_listings(driver, new_links, state_code)

        if state_data:
            print(f"\n ?? Sending {len(state_data)} listings for {state_code} to Zapier...")
            success = send_to_zapier(state_data)
            if success:
                print(f" ? All records inserted into Zapier table for {state_code}")
            else:
                print(f" ?? Some records may not have been inserted")
        else:
            print(f" ?? No valid listings extracted for {state_code}")

        if url_entries:
            append_urls_to_file(url_entries)

        print(f"\n ? {state_code} COMPLETE!\n")
        return driver

    def scrape_all_states(self, specific_states=None):
        states_to_scrape = specific_states if specific_states else list(self.states.keys())

        print(f"\n{'='*70}")
        print(f"ZILLOW ACTIVE LISTINGS SCRAPER â€” UNDETECTED CHROME")
        print(f"States          : {len(states_to_scrape)}")
        print(f"Mode            : FOR SALE (active listings)")
        print(f"Headless        : {self.headless}")
        print(f"Zapier webhook  : {ZAPIER_WEBHOOK_URL}")
        print(f"Tracking file   : {os.path.abspath(URL_TRACKING_FILE)}")
        print(f"ZPIDs tracked   : {len(self.scraped_urls)}")
        print(f"{'='*70}\n")

        driver = self._create_driver()

        for idx, state_code in enumerate(states_to_scrape, 1):
            print(f"\n{'='*70}")
            print(f"STATE {idx}/{len(states_to_scrape)}: {state_code}")
            print(f"{'='*70}\n")

            state_done    = False
            restart_count = 0

            while not state_done:
                try:
                    driver = self.scrape_state(state_code, driver)
                    state_done = True
                except Exception as e:
                    restart_count += 1
                    print(f"\n{'!'*70}")
                    print(f"  ?? {state_code} failed (restart #{restart_count}): {e}")
                    print(f"{'!'*70}\n")
                    try: driver.quit()
                    except Exception: pass
                    wait = min(15 + restart_count * 10, 60)
                    print(f"  ? Waiting {wait}s before restart...")
                    time.sleep(wait)
                    driver = self._create_driver()
                    print(f"  ? Fresh browser launched, retrying {state_code}...\n")

            if state_code != states_to_scrape[-1]:
                time.sleep(random.uniform(10, 20))

        try: driver.quit()
        except Exception: pass

        print(f"\n{'='*70}")
        print(f"? ALL STATES COMPLETED!")
        print(f"Total ZPIDs tracked: {len(self.scraped_urls)}")
        print(f"{'='*70}\n")


def main():
    scraper = ZillowActiveListingsScraper(
        headless=True,
        output_dir='./zillow_data_active',
        max_captcha_retries=3,
    )
    scraper.scrape_all_states()


if __name__ == "__main__":
    main()
