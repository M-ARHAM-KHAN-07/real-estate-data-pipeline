import os
import time
import random
import logging
import requests
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook


#  CONFIGURATION
LOGIN_URL       = "https://commercialobserver.com/sign-in/"
WEBHOOK_URL     = ""
FILE_NAME       = ""
EMAIL           = ""
PASSWORD        = ""
OFFICE_CATEGORY = "office lease"


now = datetime.now()
ARCHIVE_MONTHS = [
    f"{now.year}/{now.month:02d}", # current month
]

#  LOGGING
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)



#  BROWSER
def create_driver() -> webdriver.Chrome:
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(30)
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"},
    )
    return driver


def login(driver: webdriver.Chrome) -> None:
    log.info("Opening login page …")
    driver.get(LOGIN_URL)
    wait = WebDriverWait(driver, 20)

    # Dismiss cookie banner if present
    try:
        btn = WebDriverWait(driver, 6).until(
            EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler"))
        )
        btn.click()
        log.info("Cookie banner dismissed.")
        time.sleep(1)
    except Exception:
        pass

    # Locate email field
    email_field = None
    try:
        email_field = wait.until(EC.presence_of_element_located((By.ID, "om-user-email")))
    except Exception:
        for iframe in driver.find_elements(By.TAG_NAME, "iframe"):
            try:
                driver.switch_to.frame(iframe)
                email_field = driver.find_element(By.ID, "om-user-email")
                break
            except Exception:
                driver.switch_to.default_content()

    if email_field is None:
        raise RuntimeError("Could not find login form.")

    log.info("Entering credentials …")
    email_field.clear()
    email_field.send_keys(EMAIL)

    pwd = driver.find_element(By.ID, "om-user-pass")
    pwd.clear()
    pwd.send_keys(PASSWORD)

    submit = driver.find_element(By.CLASS_NAME, "om-login-submit")
    driver.execute_script("arguments[0].scrollIntoView(true);", submit)
    time.sleep(0.5)
    driver.execute_script("arguments[0].click();", submit)
    driver.switch_to.default_content()

    time.sleep(5)
    log.info("Login complete.")



#  URL TRACKING
def get_existing_urls() -> set:
    if not os.path.exists(FILE_NAME):
        return set()
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    urls = {row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
    log.info(f"Loaded {len(urls)} already-scraped URLs from '{FILE_NAME}'.")
    return urls


def mark_url_done(url: str) -> None:
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.append(["url", "scraped_at"])
    else:
        wb = load_workbook(FILE_NAME)
        ws = wb.active
    ws.append([url, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    wb.save(FILE_NAME)



#  ARCHIVE PAGE → article links
def get_article_links(driver: webdriver.Chrome, archive_url: str) -> list:
    log.info(f"Loading archive: {archive_url}")
    driver.get(archive_url)
    time.sleep(3)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    links = [el["href"] for el in soup.select("p.entry-title a") if el.get("href")]
    log.info(f"  Found {len(links)} article links.")
    return links



#  CORE FILTER — Office Leases category only
def is_office_lease_article(soup: BeautifulSoup) -> bool:
    """
    Go into each article page and check ONLY the article-level
    category label for the 'Office Leases' tag.

    Strategy:
      1. Strip global site chrome (nav/header/footer/aside) which
         contains the 'Office Leases' nav link on EVERY page —
         this was causing false positives before.
      2. Search known article-category CSS selectors first.
      3. Fallback: search the entire remaining (chrome-stripped) page.
    """

    
    for tag in soup.find_all(["nav", "header", "footer", "aside"]):
        tag.decompose()

    # Also remove by common ID / class names used for global chrome
    for tag in soup.find_all(
        True,
        {"id": ["menu", "navigation", "nav", "sidebar", "footer", "site-header"]},
    ):
        tag.decompose()
    for tag in soup.find_all(
        True,
        {"class": ["site-nav", "main-nav", "nav-menu", "site-header",
                   "sidebar", "widget-area", "footer-nav", "top-bar"]},
    ):
        tag.decompose()

   
    CATEGORY_SELECTORS = [
        "div.leases-label",       
        "div.channel",
        "div.article-category",
        "div.entry-category",
        "ul.post-categories",
        "span.category",
        "p.category",
        "div.label",
        "div.categories",
        "div.tags",
        "div.article-tags",
        "div.breadcrumb",
    ]

    for selector in CATEGORY_SELECTORS:
        container = soup.select_one(selector)
        if container:
            for a in container.find_all("a", href=True):
                href = a["href"].lower().rstrip("/")
                text = a.get_text(strip=True).lower()
                if href.endswith("/leases/office") and text == "office leases":
                    return True

    
    for a in soup.find_all("a", href=True):
        href = a["href"].lower().rstrip("/")
        text = a.get_text(strip=True).lower()
        if href.endswith("/leases/office") and text == "office leases":
            return True

    return False



def scrape_article(driver: webdriver.Chrome, url: str, retries: int = 2) -> tuple:
    """
    Visits the article URL.
    Returns (sub_headline, caption, body_text, is_office_lease).
    is_office_lease=False  →  caller should skip this article.
    """
    for attempt in range(retries + 1):
        try:
            driver.get(url)
            time.sleep(random.uniform(2, 3))

            # Parse once for the category check
            soup_check = BeautifulSoup(driver.page_source, "html.parser")

            if not is_office_lease_article(soup_check):
                return None, None, None, False

            # Parse a FRESH copy for content extraction
            soup = BeautifulSoup(driver.page_source, "html.parser")

            # Sub-headline
            tag = soup.find("h2", class_="sub-headline")
            sub_headline = tag.get_text(strip=True) if tag else None

            # Caption 
            caption = None
            cap_tag = soup.find("span", class_="caption")
            if cap_tag:
                credit = cap_tag.find("span", class_="media-credit")
                if credit:
                    credit.extract()
                caption = cap_tag.get_text(strip=True) or None

            # Body text
            body_text = None
            content = soup.find("div", class_="content")
            if content:
                for junk in content.find_all("om-see-also"):
                    junk.decompose()
                for junk in content.find_all(class_="see-also-related-post"):
                    junk.decompose()
                paras = [p.get_text(strip=True) for p in content.find_all("p")]
                body_text = "\n\n".join(p for p in paras if p) or None

            return sub_headline, caption, body_text, True

        except TimeoutException:
            log.warning(f"  Timeout (attempt {attempt + 1}) — {url}")
            try:
                driver.execute_script("window.stop();")
            except Exception:
                pass
            if attempt < retries:
                time.sleep(3)
            else:
                log.error(f"  Giving up after {retries + 1} attempts: {url}")
                return None, None, None, False

        except WebDriverException as e:
            log.warning(f"  WebDriver error (attempt {attempt + 1}): {e}")
            if attempt < retries:
                time.sleep(3)
            else:
                return None, None, None, False

        except Exception as e:
            log.error(f"  Unexpected error (attempt {attempt + 1}): {e}")
            if attempt < retries:
                time.sleep(3)
            else:
                return None, None, None, False

    return None, None, None, False



#  ZAPIER WEBHOOK
def send_to_zapier(url: str, sub_headline: str, caption: str,
                   body_text: str, retries: int = 3) -> bool:
    payload = {
        "url":          url,
        "sub_headline": sub_headline or "",
        "caption":      caption      or "",
        "body_text":    body_text    or "",
        "category":     OFFICE_CATEGORY,
        "date_time":    datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
    }
    for attempt in range(retries):
        try:
            r = requests.post(WEBHOOK_URL, data=payload, timeout=15)
            if r.status_code == 200:
                log.info(f"  ✅ Sent to Zapier: {url}")
                return True
            log.warning(f"  ⚠️  Zapier HTTP {r.status_code} (attempt {attempt + 1})")
        except requests.RequestException as e:
            log.warning(f"  ⚠️  Zapier request failed (attempt {attempt + 1}): {e}")
        time.sleep(2)
    log.error(f"  ❌ Could not reach Zapier after {retries} attempts: {url}")
    return False



def main() -> None:
    archive_urls = [
        f"https://commercialobserver.com/{month}/" for month in ARCHIVE_MONTHS
    ]
    log.info(f"Months to scrape: {archive_urls}")

    driver = create_driver()
    total_sent = total_skipped = total_failed = 0

    try:
        login(driver)
        existing_urls = get_existing_urls()

        for archive_url in archive_urls:
            log.info("=" * 60)
            log.info(f"Archive: {archive_url}")
            log.info("=" * 60)

            sent = skipped = failed = 0
            links = get_article_links(driver, archive_url)

            for link in links:

                
                if link in existing_urls:
                    log.info(f"  SKIP (already done): {link}")
                    skipped += 1
                    continue

                log.info(f"  Visiting: {link}")

                try:
                    sub_headline, caption, body_text, is_office_lease = scrape_article(
                        driver, link
                    )
                except Exception as e:
                    log.error(f"  Error scraping {link}: {e}")
                    failed += 1
                    continue

                
                if not is_office_lease:
                    log.info(f"  SKIP (not Office Leases): {link}")
                    skipped += 1
                    continue

                # Office Lease confirmed but no content found
                if sub_headline is None and body_text is None:
                    log.warning(f"  SKIP (no content): {link}")
                    failed += 1
                    continue

                # Send to Zapier
                if send_to_zapier(link, sub_headline, caption, body_text):
                    mark_url_done(link)
                    existing_urls.add(link)
                    sent += 1
                else:
                    log.warning(f"  Will retry on next run: {link}")
                    failed += 1

            log.info(
                f"\n  [{archive_url}]  "
                f"Sent: {sent}  |  Skipped: {skipped}  |  Failed: {failed}"
            )
            total_sent    += sent
            total_skipped += skipped
            total_failed  += failed

    finally:
        driver.quit()
        log.info("=" * 60)
        log.info("FINAL SUMMARY")
        log.info("=" * 60)
        log.info(f"  Total sent    : {total_sent}")
        log.info(f"  Total skipped : {total_skipped}")
        log.info(f"  Total failed  : {total_failed}")


if __name__ == "__main__":
    main()
